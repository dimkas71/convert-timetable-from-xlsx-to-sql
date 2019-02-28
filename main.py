from openpyxl import load_workbook

import re

CORRECT_PN = "00/0010"
INCORRECT_PN = "0d/0001"

CORRECT_PN_LIST = [
    "00/0012",
    "10/1012",
    "00/0010"
]

PN_PATTERN = "\d{2}\/\d{4}"

if (__name__ == '__main__'):
    wb = load_workbook(filename = "files/Лютий_2019_Оксана_values.xlsx", read_only = True)

    sheet_names = wb.sheetnames

    pattern = re.compile(PN_PATTERN)

    for sheet_name in sheet_names:
        print(sheet_name)
        ws = wb[sheet_name]
        max_row = wb[sheet_name].max_row
        max_column = wb[sheet_name].max_column

        for row_num in range(1, max_row + 1):
            for col_num in range(1, max_column + 1):
                cell = ws.cell(row = row_num, column = col_num)
                value = cell.value

                if (value is not None) and (type(value) is str):
                    print(col_num)
    