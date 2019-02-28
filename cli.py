from lib.config import Config

from openpyxl import load_workbook
import argparse, json, os

DEFAULT_OUTPUT = "sheets.json"

if (__name__ == '__main__'):
    
    parser = argparse.ArgumentParser(description= "Command line utility for parsing exlsx files structure...")
    
    parser.add_argument("-f", "--file", dest = "file", type = str, help = "An input xlsx file")
    parser.add_argument("-o", "--output", dest = "output", type = str, help = "An output json file")

    args = parser.parse_args()
    if (args.file):
        output = DEFAULT_OUTPUT
        if (args.output):
            output = args.output

        wb = load_workbook(args.file)
        config_list = []

        for sheet_name in wb.sheetnames:
            config_list.append(Config(file = os.path.abspath(args.file.encode("utf-8").decode("utf-8")), sheet=sheet_name.encode("utf-8").decode("utf-8"))) 

        with open(output, "w") as f:
            json.dump([c.__dict__ for c in config_list], f, indent=4, sort_keys=True)
            
