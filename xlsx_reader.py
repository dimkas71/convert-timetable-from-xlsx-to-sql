import json, argparse, sqlite3, datetime

from openpyxl import load_workbook

from lib.config import EmployeeInfo

if (__name__ == '__main__'):
    import calendar

    parser = argparse.ArgumentParser(description="Read xlsx files to objects")
    parser.add_argument("-f", "--file", dest = "file", type = str, help = "json file with configuration info")
    parser.add_argument("-m", "--month", dest = "month", type = int, help = "a month for data loading")
    parser.add_argument("-db", "--db", dest = "db", type = str, help = "An output database")

    

    args = parser.parse_args()

    config_list = None
    first, last = (None, None)

    if (args.file):
        with open(args.file, 'r', encoding = "utf-8") as fp:
            config_list = json.load(fp)

        if (args.month):
            first, last = calendar.monthrange(2019, args.month) 

    personnel_infos = []
    STEP = 4
    NO_VALUES_LIST = ['', None]
    if (config_list is not None):
        for config in config_list:
            wb = load_workbook(config["file"])
            ws = wb[config["sheet"]]
            
            row_num = config["row_from"]
            while (row_num <= config["row_to"]):
                pn, name, pos = ("", "", "")
                name = getattr(ws.cell(row = row_num, column = config["col_name"]), "value", "")
                pn = getattr(ws.cell(row = row_num, column = config["col_pn"]), "value", "")
                pos = getattr(ws.cell(row = row_num, column = config["col_position"]), "value", "")
                department = config["sheet"]
                timetable = {}


                for day_index in range(0, last):
                    list_times = []
                    code = getattr(ws.cell(row = row_num, column = config["col_first_day"] + day_index), "value", "")
                    hours = getattr(ws.cell(row = row_num + 1, column = config["col_first_day"] + day_index), "value", "")

                    if (not code in NO_VALUES_LIST) or (not hours in NO_VALUES_LIST):
                        list_times.append((code, hours))

                    code = getattr(ws.cell(row = row_num + 2, column = config["col_first_day"] + day_index), "value", "")
                    hours = getattr(ws.cell(row = row_num + 3, column = config["col_first_day"] + day_index), "value", "")

                    if (not code in NO_VALUES_LIST) or (not hours in NO_VALUES_LIST):
                        list_times.append((code, hours))    

                    timetable[str(day_index + 1)] = list_times

                if (pn is not ""):
                    personnel_infos.append(EmployeeInfo(pn, name, pos, department, timetable))
                else:
                    #TODO: logging info should be at this place....
                    pass    

                row_num = row_num + STEP


        
    TIME_SEPARATOR = ";"
    
    INSERT_SQL_EXPRESSION = "insert into timetable(pn, name, position, department, dt, value) values (?, ?, ?, ?, ?, ?)"
    
    con = sqlite3.connect(args.db)
    with con:
        for pi in personnel_infos:
            for k, v in pi.timetable.items():
                if len(v) != 0:
                    value = ""
                    for code, time in v:
                        if not code in NO_VALUES_LIST:
                            value = value + str(code) + " " + str( 0 if time is None else time) + ";"
                    con.execute(INSERT_SQL_EXPRESSION, (pi.personnelNum, pi.name, pi.position, pi.department, datetime.date(2019, 2, int(k)).strftime("%d.%m.%Y"), value,))

    